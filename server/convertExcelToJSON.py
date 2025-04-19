#!/usr/bin/env python3
"""
This script converts an Excel file to JSON format for use with the TimeTable Generator.
It extracts course and student data from the provided Excel file.

Usage:
    python convertExcelToJSON.py <input_excel_file> <output_json_file>
"""

import sys
import json
import openpyxl
from openpyxl.utils import get_column_letter
import os
from collections import defaultdict

def excel_to_json(excel_file, json_file):
    """Convert Excel file to JSON format"""
    try:
        # Load the workbook
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        
        # Dictionary to store course data
        courses = defaultdict(list)
        
        # Process each sheet in the workbook
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            
            # Find the max column with data
            max_col = 1
            while sheet.cell(row=1, column=max_col).value is not None:
                max_col += 1
            
            # Process each column with data
            for col in range(1, max_col):
                col_letter = get_column_letter(col)
                
                # Get course name (in first row)
                course_name = sheet.cell(row=1, column=col).value
                if course_name is None:
                    continue
                
                # Skip "SR" column if present (special column name)
                if course_name == "SR":
                    continue
                
                # Process student data in this column
                row = 2
                while True:
                    student_id = sheet.cell(row=row, column=col).value
                    if student_id is None:
                        break
                    
                    # Add student to the course
                    courses[str(course_name)].append(str(student_id))
                    row += 1
        
        # Add SR entry (empty for now, required by the C++ code)
        courses["SR"] = []
        
        # Write the JSON output
        with open(json_file, 'w') as f:
            json.dump(courses, f, indent=4)
        
        print(f"Successfully converted {excel_file} to {json_file}")
        return True
        
    except Exception as e:
        print(f"Error converting Excel to JSON: {str(e)}", file=sys.stderr)
        return False

if __name__ == "__main__":
    # Check command line arguments
    if len(sys.argv) != 3:
        print("Usage: python convertExcelToJSON.py <input_excel_file> <output_json_file>", file=sys.stderr)
        sys.exit(1)
    
    excel_file = sys.argv[1]
    json_file = sys.argv[2]
    
    # Check if input file exists
    if not os.path.exists(excel_file):
        print(f"Error: Input file '{excel_file}' does not exist", file=sys.stderr)
        sys.exit(1)
    
    # Convert the file
    if not excel_to_json(excel_file, json_file):
        sys.exit(1)